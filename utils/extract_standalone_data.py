#!/usr/bin/env python3
"""
Summary of Excel dependencies and how to eliminate them.

This script demonstrates what data the Python model needs from Excel
and how to extract it for standalone operation.
"""
import openpyxl
import pandas as pd
import numpy as np
import json

def extract_all_required_data(xlsx_path, output_dir="."):
    """
    Extract all data needed to run Python model without Excel.
    
    Required Data Categories:
    1. System Configuration Parameters
    2. Hourly Generation Profiles (solar, wind)
    3. Hourly Load Profile
    4. Financial Parameters (for LCOE)
    """
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_sum = wb['Summary']
    ws_hd = wb[' Hourly Data']
    ws_solar = wb['AP Solar Hourly']
    ws_wind = wb['AP Wind Hourly']
    
    print("="*70)
    print("EXTRACTING ALL DATA REQUIRED FOR STANDALONE PYTHON SIMULATION")
    print("="*70)
    
    # ========================================================================
    # 1. SYSTEM CONFIGURATION PARAMETERS
    # ========================================================================
    print("\n1. SYSTEM CONFIGURATION PARAMETERS")
    print("-" * 70)
    
    config = {
        # Load & Grid
        "total_load_mw": float(ws_sum['B8'].value or 800.0),
        "existing_solar_mwp": float(ws_sum['B10'].value or 0.0),
        "contracted_demand_mw": float(ws_hd['B4'].value or 800.0),
        "evac_limit_mw": float(ws_hd['B3'].value or 800.0),
        "tl_loss": float(ws_sum['S9'].value or 0.0),
        "wheeling_loss": float(ws_sum['S10'].value or 0.0),
        
        # Solar
        "dc_ac_ratio": float(ws_sum['B13'].value or 1.4),
        "solar_ac_mw": float(ws_sum['B14'].value or 830.0),
        "solar_degrad": float(ws_sum['B19'].value or 0.005),
        
        # Wind
        "wind_wtg_count": int(ws_sum['B21'].value or 0),
        "wind_capacity_per_wtg": float(ws_sum['B22'].value or 0.0),
        "wind_expected_cuf": float(ws_sum['B23'].value or 0.32),
        "wind_reference_cuf": float(ws_sum['B24'].value or 0.355587),
        "wind_degrad": float(ws_sum['B25'].value or 0.005),
        
        # BESS
        "bess_power_mw": float(ws_hd['B5'].value or 50.0),
        "bess_energy_mwh": float(ws_hd['B8'].value or 200.0),
        "one_way_eff": float(ws_hd['B7'].value or 0.9487),
        "bess_mode": "PV_SHIFT",  # or "FDRE"
        "soc_start_gwh": 0.0,
        
        # Other
        "p_multiplier": 1.0,  # P50/P75/P90
        "banking_enabled": True,
    }
    
    for key, value in config.items():
        print(f"  {key:<30} = {value}")
    
    # Save config
    with open(f"{output_dir}/system_config.json", 'w') as f:
        json.dump(config, f, indent=2)
    print(f"\n✓ Saved to: {output_dir}/system_config.json")
    
    # ========================================================================
    # 2. HOURLY GENERATION PROFILES (8760 hours)
    # ========================================================================
    print("\n2. HOURLY GENERATION PROFILES")
    print("-" * 70)
    
    # Solar profile (per MW AC capacity)
    solar_col = 8 if abs(config['dc_ac_ratio'] - 1.4) < 1e-9 else 9
    solar_profile = []
    hour_of_day = []  # Extract hour column from Excel
    p_multiplier = config.get('p_multiplier', 1.0)
    for i in range(1, 8761):
        val = ws_solar.cell(5+i, solar_col).value
        # Match hourly_sim_skeleton: apply p_multiplier and clip negative to 0
        solar_val = max(float(val) * p_multiplier, 0.0) if val else 0.0
        solar_profile.append(solar_val)
        # Get hour from Hourly Data sheet column G
        hour_val = ws_hd.cell(8+i, 7).value
        hour_of_day.append(int(hour_val) if hour_val is not None else i % 24)
    
    print(f"  Solar profile: {len(solar_profile)} hours")
    print(f"  Sum: {sum(solar_profile):,.0f} kWh/MW")
    print(f"  Min: {min(solar_profile):.2f}, Max: {max(solar_profile):.2f}")
    
    # Wind profile (per WTG)
    wind_profile = []
    for i in range(1, 8761):
        val = ws_wind.cell(6+i, 24).value  # Column X
        wind_profile.append(float(val) if val else 0.0)
    
    print(f"  Wind profile: {len(wind_profile)} hours")
    print(f"  Sum: {sum(wind_profile):,.0f} kWh/WTG")
    print(f"  Min: {min(wind_profile):.2f}, Max: {max(wind_profile):.2f}")
    
    # Save profiles
    df_profiles = pd.DataFrame({
        'hour_index': range(8760),
        'hour_of_day': hour_of_day,
        'solar_kwh_per_mw': solar_profile,
        'wind_kwh_per_wtg': wind_profile,
    })
    df_profiles.to_csv(f"{output_dir}/generation_profiles.csv", index=False)
    print(f"\n✓ Saved to: {output_dir}/generation_profiles.csv")
    
    # ========================================================================
    # 3. HOURLY LOAD PROFILE (8760 hours)
    # ========================================================================
    print("\n3. HOURLY LOAD PROFILE")
    print("-" * 70)
    
    load_profile = []
    for i in range(1, 8761):
        val = ws_hd.cell(8+i, 13).value  # Column M (Total Load)
        load_profile.append(float(val) if val else 0.0)
    
    print(f"  Load profile: {len(load_profile)} hours")
    print(f"  Sum: {sum(load_profile):,.0f} MWh")
    print(f"  Min: {min(load_profile):.2f}, Max: {max(load_profile):.2f}")
    
    df_load = pd.DataFrame({
        'hour': range(8760),
        'load_mw': load_profile,
    })
    df_load.to_csv(f"{output_dir}/load_profile.csv", index=False)
    print(f"\n✓ Saved to: {output_dir}/load_profile.csv")
    
    # ========================================================================
    # 4. FINANCIAL PARAMETERS (for LCOE)
    # ========================================================================
    print("\n4. FINANCIAL PARAMETERS (for LCOE)")
    print("-" * 70)
    
    try:
        ws_lcoe = wb['LCOE']
        financial = {
            "tax_rate": float(ws_lcoe['C8'].value or 0.2782),
            "discount_rate": float(ws_lcoe['C9'].value or 0.075),
            "equity_percent": float(ws_lcoe['C10'].value or 0.30),
            "loan_percent": float(ws_lcoe['C11'].value or 0.70),
            "loan_interest_rate": float(ws_lcoe['C12'].value or 0.10),
            "loan_term_years": int(ws_lcoe['C13'].value or 10),
            "project_life_years": int(ws_lcoe['C14'].value or 25),
            
            # Cost parameters (per unit)
            "solar_cost_per_mw": float(ws_lcoe['C17'].value or 40000000),  # ₹/MW
            "wind_cost_per_mw": float(ws_lcoe['C18'].value or 50000000),
            "bess_power_cost_per_mw": float(ws_lcoe['C19'].value or 5000000),
            "bess_energy_cost_per_mwh": float(ws_lcoe['C20'].value or 10000000),
            
            # Operating costs
            "solar_om_percent": float(ws_lcoe['C23'].value or 0.015),
            "wind_om_percent": float(ws_lcoe['C24'].value or 0.020),
            "bess_om_percent": float(ws_lcoe['C25'].value or 0.010),
            "insurance_percent": float(ws_lcoe['C26'].value or 0.005),
        }
        
        for key, value in financial.items():
            print(f"  {key:<30} = {value}")
        
        # Save financial params
        with open(f"{output_dir}/financial_params.json", 'w') as f:
            json.dump(financial, f, indent=2)
        print(f"\n✓ Saved to: {output_dir}/financial_params.json")
        
    except Exception as e:
        print(f"  Warning: Could not extract LCOE parameters: {e}")
    
    # ========================================================================
    # SUMMARY
    # ========================================================================
    print("\n" + "="*70)
    print("EXTRACTION COMPLETE")
    print("="*70)
    print("\nFiles created:")
    print(f"  1. system_config.json       - System configuration parameters")
    print(f"  2. generation_profiles.csv  - Solar/wind hourly profiles (8760 hrs)")
    print(f"  3. load_profile.csv         - Load hourly profile (8760 hrs)")
    print(f"  4. financial_params.json    - LCOE calculation parameters")
    print("\nThe Python model can now run standalone using these files!")
    print("="*70)

if __name__ == "__main__":
    import sys
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "oa_hybrid.xlsx"
    extract_all_required_data(xlsx_path, output_dir="standalone_data")

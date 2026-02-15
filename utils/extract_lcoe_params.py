#!/usr/bin/env python3
"""
Extract LCOE calculation parameters from the Excel workbook.
"""
import openpyxl
import json
import sys

def extract_lcoe_params(xlsx_path):
    """Extract all LCOE-relevant parameters from Excel workbook."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)  # Read cached values
    ws_lcoe = wb['LCOE']
    ws_summary = wb['Summary']
    
    # Direct LCOE sheet parameters
    params = {
        # Tax and financial rates
        'tax_rate': ws_lcoe['D9'].value,  # TR = 0.2782
        'system_degradation': ws_lcoe['D10'].value,  # SDR = 0.005
        'inflation': ws_lcoe['D13'].value,  # 0.05
        
        # Loan structure
        'equity_fraction': ws_lcoe['B14'].value,
        'loan_fraction': ws_lcoe['B15'].value,
        'loan_term_years': ws_lcoe['B16'].value,  # 10 years
        
        # CERC method params
        'post_tax_roe': ws_lcoe['B32'].value,
        'grossed_up_roe': ws_lcoe['B34'].value,
        'wacc': ws_lcoe['B35'].value,
        
        # From Summary sheet (need to lookup)
        'project_cost': ws_summary['O10'].value if ws_summary['O10'].value else None,
        'discount_rate': ws_summary['P25'].value if ws_summary['P25'].value else None,
        'interest_rate': ws_summary['P23'].value if ws_summary['P23'].value else None,
        'annual_yield_gwh': ws_summary['I11'].value if ws_summary['I11'].value else None,
        'annual_costs': ws_summary['O18'].value if ws_summary['O18'].value else None,
    }
    
    # Cost parameters from Summary sheet - scan for capital cost items
    print("\nScanning Summary sheet for cost parameters...")
    cost_params = {}
    for r in range(1, 80):
        label = ws_summary.cell(r, 1).value  # Column A
        if isinstance(label, str):
            label_clean = label.strip().lower()
            value_cell = ws_summary.cell(r, 2)  # Column B
            unit_cell = ws_summary.cell(r, 3)  # Column C
            
            # Look for cost-related parameters
            if any(keyword in label_clean for keyword in ['cost', 'capex', 'opex', 'price', 'o&m', 'tariff']):
                cost_params[f'row_{r}_{label.strip()}'] = {
                    'value': value_cell.value,
                    'unit': unit_cell.value if unit_cell.value else '',
                    'formula': isinstance(value_cell.value, str) and value_cell.value.startswith('=')
                }
    
    params['cost_items'] = cost_params
    
    # Project lifetime
    # LCOE calculation uses columns H through AF (years 1-25)
    params['project_lifetime_years'] = 25
    
    return params

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_lcoe_params.py <xlsx_path> [output_json]")
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'lcoe_params.json'
    
    params = extract_lcoe_params(xlsx_path)
    
    # Print key parameters
    print("\n" + "="*70)
    print("LCOE PARAMETERS EXTRACTED")
    print("="*70)
    print(f"\nFinancial Parameters:")
    print(f"  Tax Rate: {params['tax_rate']:.2%}")
    print(f"  System Degradation: {params['system_degradation']:.2%}")
    print(f"  Discount Rate: {params['discount_rate']}")
    print(f"  Interest Rate: {params['interest_rate']}")
    print(f"  Post-tax ROE: {params['post_tax_roe']}")
    print(f"  WACC: {params['wacc']}")
    
    print(f"\nLoan Structure:")
    print(f"  Equity Fraction: {params['equity_fraction']:.1%}")
    print(f"  Loan Fraction: {params['loan_fraction']:.1%}")
    print(f"  Loan Term: {params['loan_term_years']} years")
    
    print(f"\nProject Parameters:")
    print(f"  Project Cost: {params['project_cost']}")
    print(f"  Annual Yield: {params['annual_yield_gwh']} GWh")
    print(f"  Annual Costs: {params['annual_costs']}")
    print(f"  Lifetime: {params['project_lifetime_years']} years")
    
    print(f"\nCost Items Found: {len(params['cost_items'])}")
    print("  (See JSON for full details)")
    
    # Save to JSON
    with open(output_path, 'w') as f:
        json.dump(params, f, indent=2, default=str)
    
    print(f"\nFull parameters saved to: {output_path}")

if __name__ == '__main__':
    main()

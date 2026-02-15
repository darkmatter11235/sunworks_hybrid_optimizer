#!/usr/bin/env python3
"""
Extract a lightweight "spec" from the Excel model:
- key Summary inputs (label -> cell -> value/formula)
- ToD schedule definition ranges
- canonical Hourly Data formulas (row 9) for the main columns and BESS logic
- constants used by Hourly Data (probability multiplier, degradation, etc.)

Usage:
  python extract_excel_spec.py "AP WSB Hybrid OA Capacity Model 2.0.xlsx" spec.json
"""
from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

def is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")

def cell_value(ws, addr: str) -> Any:
    return ws[addr].value

def extract_summary_inputs(ws) -> List[Dict[str, Any]]:
    """
    Summary sheet uses a stable "label in col A, value in col B, units in col C/D" layout.
    We'll scan rows 4..40 for labels.
    """
    items = []
    for r in range(1, 80):
        label = ws.cell(r, 1).value
        if isinstance(label, str) and label.strip():
            b = ws.cell(r, 2)
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            items.append({
                "row": r,
                "label": label.strip(),
                "value_cell": b.coordinate,
                "value": b.value,
                "units": c,
                "notes": d,
                "is_formula": is_formula(b.value),
            })
    return items

def extract_tod_schedule(ws_bs) -> Dict[str, Any]:
    """
    Banking Settlement has a ToD mapping block at A2:I7 and a derived 'Monthly ToD Wise Net Delivered'
    hour boundaries in B17:C22.
    """
    # New ToD (A2:C7): label in A, from in B, to in C
    new_tod = []
    for r in range(2, 8):
        new_tod.append({
            "tod": ws_bs.cell(r, 1).value,
            "from_hour": ws_bs.cell(r, 2).value,
            "to_hour": ws_bs.cell(r, 3).value,
        })
    # Hour boundaries used by Hourly Data ToD tagging: rows 17..22 col B/C
    # These are often array-formulas; we keep the raw cell values.
    tod_bounds = []
    for r in range(17, 23):
        tod_bounds.append({
            "row": r,
            "tod_label_cell": f"A{r}",
            "from_cell": f"B{r}",
            "to_cell": f"C{r}",
            "tod_label": ws_bs[f"A{r}"].value,
            "from_expr": ws_bs[f"B{r}"].value,
            "to_expr": ws_bs[f"C{r}"].value,
        })
    return {"new_tod_block_A2C7": new_tod, "hour_boundaries_B17C22": tod_bounds}

def extract_hourly_formulas(ws_hd) -> Dict[str, Any]:
    """
    Hourly Data: take row 9 as the canonical formula template (filled down).
    We'll extract the formulas for columns we care about.
    """
    # Columns of interest (by letter)
    cols = [
        # resource and load
        "J","K","M","N","O","Q","R","S",
        # allotment
        "U","V","W","X","Y","Z","AA","AD","AE","AF","AG",
        # PV-shift BESS
        "AI","AJ","AK","AL","AM","AN",
        # FDRE BESS
        "AP","AR","AS","AT","AU","AV","AW",
        # ToD splits used by settlement
        "AZ","BA","BB","BD","BE","BF","BH","BI","BJ",
    ]
    out = {}
    for col in cols:
        addr = f"{col}9"
        out[col] = ws_hd[addr].value
    # also pull constants block B2:B8 and F4:F6
    constants = {ws_hd.cell(r,1).value: ws_hd.cell(r,2).value for r in range(2,9)}
    extra = {
        "probability_multiplier_cell": "F4",
        "probability_multiplier_expr": ws_hd["F4"].value,
        "solar_degrad_cell": "F5",
        "solar_degrad_expr": ws_hd["F5"].value,
        "wind_degrad_cell": "F6",
        "wind_degrad_expr": ws_hd["F6"].value,
    }
    return {"row9_formulas": out, "constants_B2B8": constants, "extra": extra}

def main():
    if len(sys.argv) < 3:
        print("Usage: python extract_excel_spec.py <xlsx_path> <out_json>")
        raise SystemExit(2)

    xlsx_path, out_json = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    ws_summary = wb["Summary"]
    ws_bs = wb["Banking Settlement"]
    ws_hd = wb[" Hourly Data"]

    spec = {
        "workbook": xlsx_path,
        "sheets": wb.sheetnames,
        "summary_inputs": extract_summary_inputs(ws_summary),
        "tod_schedule": extract_tod_schedule(ws_bs),
        "hourly_data": extract_hourly_formulas(ws_hd),
    }

    with open(out_json, "w") as f:
        json.dump(spec, f, indent=2, default=str)

    print(f"Wrote spec to: {out_json}")

if __name__ == "__main__":
    main()

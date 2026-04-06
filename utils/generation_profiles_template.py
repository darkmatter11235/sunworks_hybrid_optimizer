#!/usr/bin/env python3
"""Create and convert a multi-ratio generation profile Excel template."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

SOLAR_PREFIX = "solar_kwh_per_mw_dcac_"
HOURS_PER_YEAR = 8760


def _normalize_ratio(ratio: float) -> str:
    return f"{ratio:.6f}".rstrip("0").rstrip(".")


def _ratio_filename_token(ratio: float) -> str:
    return _normalize_ratio(ratio).replace(".", "_")


def _build_template_dataframe(dc_ac_ratios: Iterable[float]) -> pd.DataFrame:
    hour_index = np.arange(HOURS_PER_YEAR)
    df = pd.DataFrame(
        {
            "hour_index": hour_index,
            "hour_of_day": hour_index % 24,
            "wind_kwh_per_wtg": np.zeros(HOURS_PER_YEAR),
        }
    )
    for ratio in sorted(dc_ac_ratios):
        df[f"{SOLAR_PREFIX}{_normalize_ratio(ratio)}"] = np.zeros(HOURS_PER_YEAR)
    return df


_FILL_BLUE   = PatternFill("solid", fgColor="1F4E79")   # dark blue — section headers
_FILL_GREY   = PatternFill("solid", fgColor="D9E1F2")   # light blue-grey — field labels
_FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")   # light yellow — user-editable cells
_FONT_WHITE  = Font(bold=True, color="FFFFFF", size=11)
_FONT_BOLD   = Font(bold=True, size=10)
_FONT_NORMAL = Font(size=10)
_FONT_ITALIC = Font(italic=True, size=9, color="595959")

_LOCATION_FIELDS = ("plant_name", "place_name", "latitude", "longitude", "state_or_region", "country")


def _sec(ws, text: str) -> None:
    """Append a section-header row (dark blue, white bold text, spanning cols A–C)."""
    ws.append([text])
    row = ws.max_row
    for col in range(1, 4):
        cell = ws.cell(row, col)
        cell.fill = _FILL_BLUE
        cell.font = _FONT_WHITE
    ws.cell(row, 1).alignment = Alignment(wrap_text=False)


def _note(ws, text: str) -> None:
    """Append a single italic note row."""
    ws.append([text])
    ws.cell(ws.max_row, 1).font = _FONT_ITALIC


def _field(ws, label: str, value, hint: str = "") -> None:
    """Append a labelled input row: grey label | yellow editable cell | hint."""
    ws.append([label, value, hint])
    row = ws.max_row
    ws.cell(row, 1).fill = _FILL_GREY
    ws.cell(row, 1).font = _FONT_BOLD
    ws.cell(row, 2).fill = _FILL_YELLOW
    ws.cell(row, 2).font = _FONT_NORMAL
    ws.cell(row, 3).font = _FONT_ITALIC


def create_template(output_path: Path, dc_ac_ratios: List[float]) -> Path:
    if not dc_ac_ratios:
        raise ValueError("At least one dc_ac_ratio is required.")

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_hourly = wb.create_sheet("hourly_profiles")

    # ── column widths ────────────────────────────────────────────────────────
    ws_meta.column_dimensions["A"].width = 24
    ws_meta.column_dimensions["B"].width = 32
    ws_meta.column_dimensions["C"].width = 48

    # ── title ─────────────────────────────────────────────────────────────────
    ws_meta.append(["Hybrid RE Generation Profile Template"])
    ws_meta.cell(1, 1).font = Font(bold=True, size=14, color="1F4E79")
    ws_meta.append([])

    ws_meta.append(["This workbook is the input template for the Sunworks Hybrid Optimizer."])
    ws_meta.cell(ws_meta.max_row, 1).font = _FONT_ITALIC
    ws_meta.append(["Fill the Site Location section and the hourly_profiles sheet, then upload back into the app."])
    ws_meta.cell(ws_meta.max_row, 1).font = _FONT_ITALIC
    ws_meta.append([])

    # ── section: about ────────────────────────────────────────────────────────
    _sec(ws_meta, "ABOUT THIS TEMPLATE")
    ws_meta.append(["Sheet", "Purpose", "Action required"])
    for col in range(1, 4):
        c = ws_meta.cell(ws_meta.max_row, col)
        c.font = _FONT_BOLD
        c.fill = PatternFill("solid", fgColor="BDD7EE")
    ws_meta.append(["meta (this sheet)", "Instructions and site location input", "Fill Site Location section below"])
    ws_meta.append(["hourly_profiles", f"8 760 hourly rows (one per hour of the year)", "Fill all data columns"])
    ws_meta.append([])

    # ── section: how to fill ──────────────────────────────────────────────────
    _sec(ws_meta, "HOW TO FILL hourly_profiles")
    steps = [
        ("1", "The sheet must have exactly 8 760 rows of data (one per hour, Jan 1 00:00 → Dec 31 23:00)."),
        ("2", "Do NOT rename or delete the required columns: hour_index, hour_of_day, wind_kwh_per_wtg."),
        ("3", "hour_index must run 0 – 8 759 (integer). hour_of_day must be 0 – 23 repeating."),
        ("4", f"For each DC/AC ratio you want to model, add a solar column named:  {SOLAR_PREFIX}<ratio>"),
        ("5", f"Example column names:  {SOLAR_PREFIX}1.4   and   {SOLAR_PREFIX}1.45"),
        ("6", "Solar values are in kWh per MW AC per hour (i.e. specific yield, hourly basis)."),
        ("7", "Wind values are in kWh per WTG per hour."),
        ("8", "Zero-fill any hours with no generation. Do not leave cells blank in data rows."),
    ]
    for num, text in steps:
        ws_meta.append([f"Step {num}", text])
        ws_meta.cell(ws_meta.max_row, 1).font = _FONT_BOLD
        ws_meta.cell(ws_meta.max_row, 2).font = _FONT_NORMAL
        ws_meta.cell(ws_meta.max_row, 2).alignment = Alignment(wrap_text=True)
    ws_meta.row_dimensions[ws_meta.max_row].height = 28
    ws_meta.append([])

    # ── section: column reference ─────────────────────────────────────────────
    _sec(ws_meta, "COLUMN REFERENCE — hourly_profiles")
    ws_meta.append(["Column name", "Required", "Unit", "Description"])
    for col in range(1, 5):
        c = ws_meta.cell(ws_meta.max_row, col)
        c.font = _FONT_BOLD
        c.fill = PatternFill("solid", fgColor="BDD7EE")
    ws_meta.column_dimensions["D"].width = 48
    col_ref = [
        ("hour_index",                  "Yes", "integer 0–8759", "Sequential hour number for the year"),
        ("hour_of_day",                 "Yes", "integer 0–23",   "Hour within the day (0 = midnight)"),
        ("wind_kwh_per_wtg",            "Yes", "kWh / WTG",      "Wind energy per turbine per hour"),
        (f"{SOLAR_PREFIX}1.4",  "Per ratio", "kWh / MW AC",  "Solar specific yield for DC/AC ratio 1.4"),
        (f"{SOLAR_PREFIX}1.45", "Per ratio", "kWh / MW AC",  "Solar specific yield for DC/AC ratio 1.45"),
        ("(add more solar columns as needed)", "", "", ""),
    ]
    for row_data in col_ref:
        ws_meta.append(list(row_data))
        ws_meta.cell(ws_meta.max_row, 1).font = Font(bold=True, size=9, color="1F4E79")
        for c in range(2, 5):
            ws_meta.cell(ws_meta.max_row, c).font = _FONT_NORMAL
    ws_meta.append([])

    # ── section: site location ────────────────────────────────────────────────
    _sec(ws_meta, "SITE LOCATION   ← fill the yellow cells")
    _note(ws_meta, "Enter location data here. This is saved alongside the generation profiles.")
    ws_meta.append([])
    _field(ws_meta, "plant_name",     "", "Plant / project name  (optional)")
    _field(ws_meta, "place_name",     "", "Village, town or site name  (required)")
    _field(ws_meta, "latitude",       "", "Decimal degrees, −90 to 90  (required, e.g. 14.6819)")
    _field(ws_meta, "longitude",      "", "Decimal degrees, −180 to 180  (required, e.g. 77.6006)")
    _field(ws_meta, "state_or_region","", "State or region  (optional)")
    _field(ws_meta, "country",        "",  "Country  (optional)")
    ws_meta.append([])

    # ── hourly_profiles sheet ─────────────────────────────────────────────────
    df_template = _build_template_dataframe(dc_ac_ratios)
    for row in dataframe_to_rows(df_template, index=False, header=True):
        ws_hourly.append(row)
    ws_hourly.freeze_panes = "A2"

    # Bold the header row
    for cell in ws_hourly[1]:
        cell.font = _FONT_BOLD
        cell.fill = PatternFill("solid", fgColor="BDD7EE")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _read_plant_info(template_path: Path) -> Dict[str, object]:
    """Read location fields from the meta sheet (col A = field name, col B = value)."""
    try:
        ws_data = pd.read_excel(template_path, sheet_name="meta", header=None)
    except ValueError:
        return {}

    meta: Dict[str, object] = {}
    for _, row in ws_data.iterrows():
        if len(row) < 2:
            continue
        field = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if field not in _LOCATION_FIELDS:
            continue
        value = row.iloc[1]
        if pd.isna(value) or str(value).strip() == "":
            continue
        meta[field] = value

    # Fall back to legacy plant_info sheet if meta had nothing
    if not meta:
        try:
            df_plant = pd.read_excel(template_path, sheet_name="plant_info")
            if {"field", "value"}.issubset(df_plant.columns):
                for _, row in df_plant.iterrows():
                    field = str(row.get("field", "")).strip()
                    if field not in _LOCATION_FIELDS:
                        continue
                    value = row.get("value")
                    if pd.isna(value):
                        continue
                    meta[field] = value
        except ValueError:
            pass

    for numeric_field in ("latitude", "longitude"):
        if numeric_field in meta:
            try:
                meta[numeric_field] = float(meta[numeric_field])
            except (TypeError, ValueError):
                raise ValueError(f"Invalid {numeric_field}: {meta[numeric_field]}")

    if "latitude" in meta and not (-90.0 <= float(meta["latitude"]) <= 90.0):
        raise ValueError("Latitude must be in range -90 to 90")
    if "longitude" in meta and not (-180.0 <= float(meta["longitude"]) <= 180.0):
        raise ValueError("Longitude must be in range -180 to 180")

    return meta


def _extract_ratio_columns(columns: Iterable[str]) -> List[Tuple[float, str]]:
    ratio_columns: List[Tuple[float, str]] = []
    for column in columns:
        if not column.startswith(SOLAR_PREFIX):
            continue

        raw_token = column[len(SOLAR_PREFIX) :].strip()
        token = raw_token.replace("_", ".")
        try:
            ratio = float(token)
        except ValueError:
            continue
        ratio_columns.append((ratio, column))

    ratio_columns.sort(key=lambda item: item[0])
    return ratio_columns


def _validate_hourly_dataframe(df_hourly: pd.DataFrame) -> None:
    required = {"hour_index", "hour_of_day", "wind_kwh_per_wtg"}
    missing = sorted(required - set(df_hourly.columns))
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")
    if len(df_hourly) != HOURS_PER_YEAR:
        raise ValueError(
            f"Expected {HOURS_PER_YEAR} rows in hourly_profiles sheet, found {len(df_hourly)}"
        )


def convert_template_to_profiles(
    template_path: Path,
    output_dir: Path,
    sheet_name: str = "hourly_profiles",
    default_ratio: float | None = None,
) -> List[Path]:
    df_hourly = pd.read_excel(template_path, sheet_name=sheet_name)
    _validate_hourly_dataframe(df_hourly)
    plant_info = _read_plant_info(template_path)

    ratio_columns = _extract_ratio_columns(df_hourly.columns)
    if not ratio_columns:
        raise ValueError(
            "No solar ratio columns found. Add columns like "
            f"{SOLAR_PREFIX}1.4, {SOLAR_PREFIX}1.45"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    written: List[Path] = []
    by_ratio: Dict[float, Path] = {}

    for ratio, solar_column in ratio_columns:
        location_name = plant_info.get("place_name") or plant_info.get("plant_name") or ""
        latitude = plant_info.get("latitude") if "latitude" in plant_info else np.nan
        longitude = plant_info.get("longitude") if "longitude" in plant_info else np.nan
        out_df = pd.DataFrame(
            {
                "hour_index": df_hourly["hour_index"].astype(int),
                "hour_of_day": df_hourly["hour_of_day"].astype(int),
                "solar_kwh_per_mw": pd.to_numeric(df_hourly[solar_column], errors="coerce").fillna(0.0),
                "dc_ac_ratio": ratio,
                "wind_kwh_per_wtg": pd.to_numeric(df_hourly["wind_kwh_per_wtg"], errors="coerce").fillna(0.0),
                "location_name": location_name,
                "latitude": latitude,
                "longitude": longitude,
            }
        )

        out_file = output_dir / f"generation_profiles_dcac_{_ratio_filename_token(ratio)}.csv"
        out_df.to_csv(out_file, index=False)
        written.append(out_file)
        by_ratio[ratio] = out_file

    if default_ratio is None:
        default_ratio = 1.4

    if default_ratio in by_ratio:
        default_source = by_ratio[default_ratio]
    else:
        default_source = written[0]

    default_target = output_dir / "generation_profiles.csv"
    pd.read_csv(default_source).to_csv(default_target, index=False)
    written.append(default_target)

    if plant_info:
        metadata_target = output_dir / "plant_location.json"
        with open(metadata_target, "w", encoding="utf-8") as fp:
            json.dump(plant_info, fp, indent=2)
        written.append(metadata_target)

    return written


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create or convert multi-dc_ac_ratio generation profile templates"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    create_parser = subparsers.add_parser("create-template", help="Create empty Excel template")
    create_parser.add_argument(
        "--output",
        default="standalone_data/generation_profiles_template.xlsx",
        help="Output template workbook path",
    )
    create_parser.add_argument(
        "--ratios",
        nargs="+",
        type=float,
        default=[1.4, 1.45],
        help="dc_ac_ratio values to include as solar columns",
    )

    convert_parser = subparsers.add_parser("convert", help="Convert template workbook to CSV profiles")
    convert_parser.add_argument("--input", required=True, help="Template workbook path")
    convert_parser.add_argument(
        "--output-dir",
        default="standalone_data",
        help="Directory for generation_profiles*.csv outputs",
    )
    convert_parser.add_argument(
        "--sheet",
        default="hourly_profiles",
        help="Worksheet name containing hourly data",
    )
    convert_parser.add_argument(
        "--default-ratio",
        type=float,
        default=1.4,
        help="Ratio to mirror into generation_profiles.csv when available",
    )

    args = parser.parse_args()

    if args.command == "create-template":
        output = create_template(Path(args.output), args.ratios)
        print(f"Template created: {output}")
        return 0

    if args.command == "convert":
        written_files = convert_template_to_profiles(
            template_path=Path(args.input),
            output_dir=Path(args.output_dir),
            sheet_name=args.sheet,
            default_ratio=args.default_ratio,
        )
        print("Wrote files:")
        for path in written_files:
            print(f"  - {path}")
        return 0

    return 1


if __name__ == "__main__":
    raise SystemExit(main())

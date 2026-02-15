import openpyxl
import numpy as np
import pandas as pd

wb = openpyxl.load_workbook('oa_hybrid.xlsx', data_only=True)
ws_lcoe = wb['LCOE']

print("="*80)
print("DETAILED EXCEL LCOE SHEET ANALYSIS")
print("="*80)

# Scan the entire sheet to understand the structure
print("\nScanning for key values...")
for row in range(1, 60):
    for col in range(1, 8):
        cell = ws_lcoe.cell(row, col)
        if cell.value:
            val_str = str(cell.value)
            # Look for keywords
            if any(keyword in val_str.lower() for keyword in [
                'lcoe', 'npv', 'cost', 'energy', 'project', 'equity', 
                'loan', 'depreciation', 'interest', 'tax', 'opex', 
                'discount', 'degradation', 'delivered', 'year 1'
            ]):
                # Get surrounding cells for context
                print(f"{cell.coordinate:5s} {val_str[:60]:60s}", end="")
                if col < 7:
                    next_cell = ws_lcoe.cell(row, col+1)
                    if next_cell.value and isinstance(next_cell.value, (int, float)):
                        print(f" → {next_cell.value}")
                    else:
                        print()
                else:
                    print()

print("\n" + "="*80)
print("YEAR-BY-YEAR DATA (if available)")
print("="*80)

# Look for year-by-year columns (typically years are in columns for NPV calculations)
# Check row 2 for year headers
print("\nChecking row 2 for year headers:")
for col in range(1, 30):
    val = ws_lcoe.cell(2, col).value
    if val and isinstance(val, (int, float)) and 0 < val < 30:
        print(f"Column {openpyxl.utils.get_column_letter(col)}: Year {val}")

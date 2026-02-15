#!/usr/bin/env python3
"""
Hourly simulation skeleton that mirrors the *Houry Data* sheet logic.

This does NOT attempt to be a general "Excel formula evaluator".
Instead, it re-implements the spreadsheet math explicitly, which is the
most maintainable approach for a Python port.

Key ideas:
- treat each hour as 1h, so MW and MWh/h are numerically interchangeable.
- compute generation -> allocate to consumption/bess/banking -> compute losses -> compute grid import
- implement BESS as a stateful SOC recursion.

You can run this directly against the workbook to load the same solar & wind profiles.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Literal, Tuple
import numpy as np
import pandas as pd
import openpyxl

BessMode = Literal["PV_SHIFT", "FDRE"]

@dataclass
class Config:
    year: int = 1

    # Load & OA
    total_load_mw: float = 500.0
    existing_solar_mwp: float = 0.0
    contracted_demand_mw: float = 0.0
    evac_limit_mw: float = 1200.0
    tl_loss: float = 0.0       # Summary!S9
    wheeling_loss: float = 0.0 # Summary!S10

    # Generation (new project)
    dc_ac_ratio: float = 1.4
    solar_ac_mw: float = 900.0
    wind_wtg_count: float = 0.0
    wind_expected_cuf: float = 0.45
    wind_reference_cuf: float = 0.342

    # Degradation
    solar_degrad: float = 0.0
    wind_degrad: float = 0.0
    p_multiplier: float = 1.0  # P50/P75/P90 multiplier

    # Banking
    banking_enabled: bool = True

    # BESS
    bess_mode: BessMode = "PV_SHIFT"
    one_way_eff: float = 0.915
    bess_power_mw: float = 50.0
    bess_energy_mwh: float = 200.0
    soc_start_gwh: float = 0.0   # Excel uses *1000 -> convert to MWh

    # FDRE parameters (if mode=FDRE)
    firm_dispatch_start_hour: int = 19
    firm_dispatch_end_hour: int = 23
    firm_power_mw: float = 800.0

    @property
    def loss_factor(self) -> float:
        return (1.0 - self.tl_loss) * (1.0 - self.wheeling_loss)

    @property
    def soc_start_mwh(self) -> float:
        return self.soc_start_gwh * 1000.0

def assign_tod(hour: np.ndarray,
               tod_breaks: Tuple[Tuple[str,int,int], ...] = (("normal",0,5),("peak",5,9),("offpeak",9,17),("normal",17,19),("peak",19,23),("normal",23,24))
               ) -> np.ndarray:
    """
    The workbook tags TOD using Banking Settlement hour breakpoints. This helper uses
    a default schedule matching Excel:
      normal:  0-4, 17-18, 23
      peak:    5-8, 19-22
      offpeak: 9-16
    Replace tod_breaks with the exact schedule you want.
    """
    out = np.empty_like(hour, dtype=object)
    for label, start, end in tod_breaks:
        if end == 0:
            mask = (hour >= start) | (hour < 24)  # (wrap)
        else:
            mask = (hour >= start) & (hour < end)
        out[mask] = label
    return out

def load_profiles_from_workbook(xlsx_path: str, cfg: Config) -> pd.DataFrame:
    """
    Reads:
      - Hourly calendar from ' Hourly Data' columns C:D (Date, Time)
      - Solar profile from 'AP Solar Hourly' (H if DC/AC=1.4 else I) (kW or kWh scale)
      - Wind profile from 'AP Wind Hourly' column X
      - Load profile from ' Hourly Data' column M (Total Load)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)  # read cached values if present
    hd = wb[" Hourly Data"]
    solar = wb["AP Solar Hourly"]
    wind = wb["AP Wind Hourly"]

    # Hourly calendar: rows 9..8768
    n = 8760
    # Column D contains complete datetime values, Column G contains hour-of-day
    times = [hd.cell(8+i, 4).value for i in range(1, n+1)]  # col D
    dt = pd.to_datetime(times)
    # Use Excel's hour column directly (col G) to avoid rounding issues at hour boundaries
    hour = np.array([hd.cell(8+i, 7).value for i in range(1, n+1)], dtype=int)
    month = dt.month.to_numpy()
    day = dt.day.to_numpy()

    # Solar profile: AP Solar Hourly row 6..8765
    solar_col = 8 if abs(cfg.dc_ac_ratio - 1.4) < 1e-9 else 9  # H=8, I=9
    solar_raw = np.array([solar.cell(5+i, solar_col).value for i in range(1, n+1)], dtype=float)
    solar_mwh_per_mw = np.maximum(solar_raw * cfg.p_multiplier, 0.0) / 1000.0  # Excel: /1000

    # Wind: AP Wind Hourly col X=24 row 7..8766 (check workbook)
    wind_raw = np.array([wind.cell(6+i, 24).value for i in range(1, n+1)], dtype=float)
    wind_mwh_per_wtg = wind_raw * (cfg.wind_expected_cuf / cfg.wind_reference_cuf) / 1000.0

    # Load profile: Hourly Data col M (Total Load)
    load_mw = np.array([hd.cell(8+i, 13).value for i in range(1, n+1)], dtype=float)

    df = pd.DataFrame({
        "dt": dt,
        "month": month,
        "day": day,
        "hour": hour,
        "tod": assign_tod(hour),
        "solar_mwh_per_mw": solar_mwh_per_mw,
        "wind_mwh_per_wtg": wind_mwh_per_wtg,
        "load_mw": load_mw,
    })
    return df

def simulate_hourly(cfg: Config, df: pd.DataFrame) -> pd.DataFrame:
    """
    Mirrors key Hourly Data sheet equations.

    Notation:
      - quantities are MWh per hour (numeric MW for 1h intervals)
      - "injection" = before OA losses
      - "delivered" = after OA losses (loss_factor)
    """
    df = df.copy()

    # Load and generation (with degradation)
    degr_s = (1.0 - cfg.solar_degrad) ** (cfg.year - 1)
    degr_w = (1.0 - cfg.wind_degrad) ** (cfg.year - 1)

    # Use load profile from dataframe if available, otherwise use constant
    if "load_mw" in df.columns:
        df["load_total"] = df["load_mw"]
    else:
        df["load_total"] = cfg.total_load_mw
    
    df["existing_solar"] = (cfg.existing_solar_mwp / cfg.dc_ac_ratio) * df["solar_mwh_per_mw"] * degr_s
    df["net_load"] = df["load_total"] - df["existing_solar"]

    df["solar_gen"] = cfg.solar_ac_mw * df["solar_mwh_per_mw"] * degr_s
    df["wind_gen"] = cfg.wind_wtg_count * df["wind_mwh_per_wtg"] * degr_w
    df["gen_total"] = df["solar_gen"] + df["wind_gen"]

    # BESS dispatch state
    soc = np.zeros(len(df) + 1, dtype=float)
    soc[0] = cfg.soc_start_mwh

    charge = np.zeros(len(df), dtype=float)
    discharge = np.zeros(len(df), dtype=float)

    # Precompute some arrays
    loss_factor = cfg.loss_factor
    hour = df["hour"].to_numpy()
    tod = df["tod"].to_numpy()
    net_load = df["net_load"].to_numpy()
    gen_total = df["gen_total"].to_numpy()

    if cfg.bess_mode == "FDRE":
        firm_energy_required = cfg.firm_power_mw * (cfg.firm_dispatch_end_hour - cfg.firm_dispatch_start_hour)
        firm_energy_required_cons = firm_energy_required / loss_factor

    # We need values of to_consumption_inj (V) to compute PV_SHIFT discharge limits (depends on evac headroom).
    # So we compute consumption first with a placeholder available. We'll refine after BESS if needed.
    # Excel effectively computes V without including BESS discharge.
    to_consumption_inj = np.minimum(net_load / loss_factor, np.minimum(gen_total, cfg.evac_limit_mw))

    for t in range(len(df)):
        # PV_SHIFT discharge logic (Excel AJ/AK/AL)
        if cfg.bess_mode == "PV_SHIFT":
            # Component AJ (Excel col AJ): discharge to cover load above contracted demand
            # Formula: MAX(0, MIN(netload_M - contracted_demand, bess_power, soc*eff, evac_limit - V))
            aj = 0.0
            if net_load[t] > cfg.contracted_demand_mw:
                aj = min(net_load[t] - cfg.contracted_demand_mw,
                        cfg.bess_power_mw,
                        soc[t] * cfg.one_way_eff,
                        max(0.0, cfg.evac_limit_mw - to_consumption_inj[t]))
                aj = max(0.0, aj)
            
            # Component AK (Excel col AK): in "peak" hours, discharge to cover deficit
            # Formula: IF(TOD="peak", MAX(0, MIN(netload_M - gen_S, bess_power-AJ, soc*eff-AJ/eff, evac_limit-V-AJ)), 0)
            ak = 0.0
            if tod[t] == "peak" and net_load[t] > gen_total[t]:
                remaining_power = cfg.bess_power_mw - aj
                remaining_soc = soc[t] - aj / cfg.one_way_eff
                remaining_evac = max(0.0, cfg.evac_limit_mw - to_consumption_inj[t] - aj)
                ak = min(net_load[t] - gen_total[t],
                        remaining_power,
                        remaining_soc * cfg.one_way_eff,
                        remaining_evac)
                ak = max(0.0, ak)
            
            d = aj + ak

            # Charging from excess generation (Excel AM)
            # Formula: MAX(0, MIN(gen_S - V, bess_power, (capacity - soc)/eff))
            excess = max(0.0, gen_total[t] - to_consumption_inj[t])
            available_capacity = (cfg.bess_energy_mwh - soc[t]) / cfg.one_way_eff
            c = min(excess, cfg.bess_power_mw, available_capacity)
            c = max(c, 0.0)

            # SOC update (Excel AN)
            soc[t+1] = min(cfg.bess_energy_mwh,
                          max(0.0, soc[t] + c * cfg.one_way_eff - d / cfg.one_way_eff))
            charge[t] = c
            discharge[t] = d

        # FDRE mode (Excel AR/AS/AT/AU/AV/AW)
        else:
            # First pass: compute daily generation sum (Excel AP = SUMIF day)
            # We need the day context; compute this beforehand for efficiency
            day_num = df.iloc[t]["day"]
            month_num = df.iloc[t]["month"]
            
            # Calculate firm energy requirement (Excel AQ)
            # =IF(SUMIFS(gen_total, month=this_month, day=this_day) >= firm_energy_required_cons, firm_energy_required_cons, 0)
            daily_mask = (df["month"] == month_num) & (df["day"] == day_num)
            daily_gen = gen_total[daily_mask].sum()
            
            firm_target = firm_energy_required_cons if daily_gen >= firm_energy_required_cons else 0.0
            
            # Discharge window (Excel AT): fixed firm power during dispatch hours
            in_window = (hour[t] >= cfg.firm_dispatch_start_hour) and (hour[t] < cfg.firm_dispatch_end_hour)
            if in_window and firm_target > 0:
                d = min(cfg.firm_power_mw, cfg.bess_power_mw, soc[t] * cfg.one_way_eff)
            else:
                d = 0.0
            
            # Charging (Excel AU): allocate firm_target across pre-window hours
            # Formula: IF(hour < dispatch_start, MIN(gen_S - V, firm_target / count_prewindow_hours, bess_power, (capacity-soc)/eff), 0)
            if hour[t] < cfg.firm_dispatch_start_hour and firm_target > 0:
                # Count pre-window hours in this day
                prewindow_hours = cfg.firm_dispatch_start_hour  # hours 0 to dispatch_start-1
                target_per_hour = firm_target / prewindow_hours if prewindow_hours > 0 else 0.0
                
                excess = max(0.0, gen_total[t] - to_consumption_inj[t])
                available_capacity = (cfg.bess_energy_mwh - soc[t]) / cfg.one_way_eff
                c = min(excess, target_per_hour, cfg.bess_power_mw, available_capacity)
                c = max(0.0, c)
            else:
                c = 0.0
            
            # SOC update (Excel AW)
            soc[t+1] = min(cfg.bess_energy_mwh,
                          max(0.0, soc[t] + c * cfg.one_way_eff - d / cfg.one_way_eff))
            charge[t] = c
            discharge[t] = d

    df["bess_charge_inj"] = charge
    df["bess_discharge_inj"] = discharge
    df["soc_mwh"] = soc[1:]

    # Allocation (Excel-style)
    if cfg.bess_mode == "FDRE":
        available_for_consumption = df["gen_total"] - df["bess_charge_inj"]
    else:
        available_for_consumption = df["gen_total"]

    df["to_consumption_inj"] = np.minimum(df["net_load"] / loss_factor,
                                          np.minimum(available_for_consumption, cfg.evac_limit_mw))
    df["excess_after_cons"] = available_for_consumption - df["to_consumption_inj"]

    df["to_battery_inj"] = df["bess_charge_inj"]

    if cfg.banking_enabled:
        df["to_banking_inj"] = np.minimum(df["excess_after_cons"] - df["to_battery_inj"],
                                          cfg.evac_limit_mw - df["to_consumption_inj"])
        df["to_banking_inj"] = df["to_banking_inj"].clip(lower=0.0)
    else:
        df["to_banking_inj"] = 0.0

    df["curtailed"] = df["gen_total"] - df["to_consumption_inj"] - df["to_battery_inj"] - df["to_banking_inj"]
    df["net_injection"] = df["to_consumption_inj"] + df["to_banking_inj"] + df["bess_discharge_inj"]

    # Delivered side (after OA losses)
    df["direct_delivered"] = df["to_consumption_inj"] * loss_factor
    df["balance_after_direct"] = df["net_load"] - df["direct_delivered"]

    if cfg.bess_mode == "FDRE":
        # TODO: replace with FDRE discharge column once ported exactly
        df["from_bess_delivered"] = np.minimum(df["bess_discharge_inj"] * loss_factor, df["balance_after_direct"])
    else:
        df["from_bess_delivered"] = np.minimum(df["bess_discharge_inj"] * loss_factor, df["balance_after_direct"])

    df["from_grid"] = df["balance_after_direct"] - df["from_bess_delivered"]

    # Helpful ToD splits for settlement (match Hourly Data columns AZ/BA/BB, BD/BE/BF, BH/BI/BJ)
    df["banked_after_losses"] = df["to_banking_inj"] * loss_factor
    df["delivered_total"] = df["direct_delivered"] + df["from_bess_delivered"]
    df["grid_tod_offpeak"] = np.where(df["tod"] == "offpeak", df["from_grid"], 0.0)
    df["grid_tod_normal"]  = np.where(df["tod"] == "normal",  df["from_grid"], 0.0)
    df["grid_tod_peak"]    = np.where(df["tod"] == "peak",    df["from_grid"], 0.0)

    return df

if __name__ == "__main__":
    # Quick smoke-run against the workbook in the current folder (edit path as needed)
    import sys
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "AP WSB Hybrid OA Capacity Model 2.0.xlsx"
    cfg = Config()
    df0 = load_profiles_from_workbook(xlsx, cfg)
    out = simulate_hourly(cfg, df0)
    print(out[["net_load","gen_total","to_consumption_inj","to_banking_inj","bess_charge_inj","bess_discharge_inj","from_grid"]].sum())

import openpyxl
import numpy as np

wb = openpyxl.load_workbook('oa_hybrid.xlsx', data_only=True)
ws_lcoe = wb['LCOE']

print("="*80)
print("EXTRACTING YEAR-BY-YEAR LCOE CALCULATIONS FROM EXCEL")
print("="*80)

# Years are in columns H (1) through AF (25)
start_col = 8  # Column H
n_years = 25

# Key NPV values
npv_generation = ws_lcoe['B40'].value  # Column B, not A
npv_costs = ws_lcoe['B41'].value
lcoe_no_ad = ws_lcoe['B20'].value

print(f"\nFinal Results from Excel:")
if npv_generation:
    print(f"NPV of Generation: ₹{float(npv_generation):.2f} Crore")
if npv_costs:
    print(f"NPV of Costs:      ₹{float(npv_costs):.2f} Crore")
if lcoe_no_ad:
    print(f"LCOE (no AD):      ₹{float(lcoe_no_ad):.4f}/kWh")

# Try to extract year-by-year values
print("\n" + "="*80)
print("YEAR-BY-YEAR BREAKDOWN")
print("="*80)

# Look for energy delivered row
print("\nSearching for energy/generation rows...")
for row in range(3, 50):
    a_val = ws_lcoe.cell(row, 1).value
    if a_val:
        a_str = str(a_val).lower()
        if any(word in a_str for word in ['energy', 'generation', 'delivered', 'kwh', 'gwh']):
            print(f"\nRow {row}: {a_val}")
            year_1_val = ws_lcoe.cell(row, start_col).value
            year_2_val = ws_lcoe.cell(row, start_col+1).value
            year_3_val = ws_lcoe.cell(row, start_col+2).value
            if year_1_val:
                print(f"  Year 1: {year_1_val}")
                print(f"  Year 2: {year_2_val}")
                print(f"  Year 3: {year_3_val}")

# Look for key cost/cash flow rows
print("\nSearching for cost/cash flow rows...")
cost_keywords = ['depreciation', 'interest', 'loan', 'opex', 'equity', 'cost', 'cash', 'payment']
for row in range(3, 50):
    a_val = ws_lcoe.cell(row, 1).value
    g_val = ws_lcoe.cell(row, 7).value  # Column G often has labels
    
    # Check column A
    if a_val:
        a_str = str(a_val).lower()
        if any(word in a_str for word in cost_keywords):
            year_1_val = ws_lcoe.cell(row, start_col).value
            if year_1_val and isinstance(year_1_val, (int, float)):
                print(f"\nRow {row} (Col A): {a_val}")
                print(f"  Year 1: {year_1_val:.2f}")
    
    # Check column G
    if g_val:
        g_str = str(g_val).lower()
        if any(word in g_str for word in cost_keywords):
            year_1_val = ws_lcoe.cell(row, start_col).value
            if year_1_val and isinstance(year_1_val, (int, float)):
                print(f"\nRow {row} (Col G): {g_val}")
                print(f"  Year 1: {year_1_val:.2f}")

# Extract system degradation rate
print("\n" + "="*80)
print("PARAMETERS")
print("="*80)
degradation = ws_lcoe['D10'].value
discount_rate = ws_lcoe['D11'].value
tax_rate = ws_lcoe['D9'].value

if degradation:
    print(f"System Degradation: {float(degradation) if degradation else 'N/A'}")
if discount_rate:
    print(f"Discount Rate:      {float(discount_rate) if discount_rate else 'N/A'}")
if tax_rate:
    print(f"Tax Rate:           {float(tax_rate) if tax_rate else 'N/A'}")

# Try to find first year energy
print("\n" + "="*80)
print("Scanning rows 3-15 for year 1 values in column H...")
print("="*80)
for row in range(3, 16):
    label_a = ws_lcoe.cell(row, 1).value
    label_g = ws_lcoe.cell(row, 7).value
    val_h = ws_lcoe.cell(row, 8).value
    
    if val_h and isinstance(val_h, (int, float)):
        print(f"Row {row:2d} H: {val_h:15.2f} | A: {str(label_a)[:30] if label_a else '':30s} | G: {str(label_g)[:30] if label_g else ''}")

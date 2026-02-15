import openpyxl
import numpy as np

wb = openpyxl.load_workbook('oa_hybrid.xlsx')
wb_val = openpyxl.load_workbook('oa_hybrid.xlsx', data_only=True)
ws = wb['LCOE']
ws_val = wb_val['LCOE']

print("="*80)
print("EXCEL NPV FORMULAS")
print("="*80)

# Check key NPV formulas
cells = ['E5', 'E7', 'E12', 'E17', 'E18']
labels = ['NPV OPEX', 'NPV Depreciation', 'NPV Energy', 'NPV Interest', 'NPV Loan Payment']

for cell, label in zip(cells, labels):
    formula = ws[cell].value
    value = ws_val[cell].value
    print(f"\n{label} ({cell}):")
    print(f"  Formula: {formula}")
    print(f"  Value:   {value:.2f}")

# Check what's in row 39 (the yearly costs that are NPVed in B41)
print("\n" + "="*80)
print("ROW 39 (Yearly Costs for NPV)")
print("="*80)

g39 = ws_val.cell(39, 7).value  # Column G
print(f"G39 (label): {g39}")

# Get first few year values
for col in range(8, 13):  # H through L (years 1-5)
    val = ws_val.cell(39, col).value
    print(f"  Year {col-7}: {val:.2f}")

# Check row 5 (Annual Costs) - year by year
print("\n" + "="*80)
print("ROW 5 (Annual Costs)")
print("="*80)

g5_label = ws_val.cell(5, 7).value
print(f"G5 (label): {g5_label}")

for col in range(8, 13):  # Years 1-5
    val = ws_val.cell(5, col).value
    print(f"  Year {col-7}: {val:.2f}")

# Check if there's inflation or escalation
print("\n" + "="*80)
print("Checking for OPEX escalation...")
print("="*80)

year_1_opex = ws_val.cell(5, 8).value
year_2_opex = ws_val.cell(5, 9).value
year_3_opex = ws_val.cell(5, 10).value

if abs(year_1_opex - year_2_opex) < 0.01:
    print("OPEX is constant (no escalation)")
else:
    escalation = (year_2_opex / year_1_opex - 1) * 100
    print(f"OPEX escalates by {escalation:.2f}% per year")

# Calculate NPV manually
discount_rate = ws_val['D11'].value
annual_opex = ws_val['D5'].value

print(f"\nDiscount rate: {discount_rate}")
print(f"Annual OPEX (D5): {annual_opex:.2f}")

# Method 1: Standard NPV (end of period)
npv_1 = sum(annual_opex / ((1 + discount_rate) ** (year + 1)) for year in range(25))
print(f"\nMethod 1 (end of period):  NPV = {npv_1:.2f}")

# Method2: Beginning of period
npv_2 = sum(annual_opex / ((1 + discount_rate) ** year) for year in range(25))
print(f"Method 2 (begin of period): NPV = {npv_2:.2f}")

# Method 3: Use actual year values from row 5
opex_values = [ws_val.cell(5, col).value for col in range(8, 33)]  # Years 1-25
npv_3 = sum(opex_values[year] / ((1 + discount_rate) ** (year + 1)) for year in range(25))
print(f"Method 3 (using row 5 values): NPV = {npv_3:.2f}")

# Method 4: Excel NPV function equivalent (starts at year 0?)
npv_4 = sum(opex_values[year] / ((1 + discount_rate) ** year) for year in range(25))
print(f"Method 4 (year 0 start): NPV = {npv_4:.2f}")

excel_npv_opex = ws_val['E5'].value
print(f"\nExcel NPV OPEX (E5): {excel_npv_opex:.2f}")

# Which method matches?
methods = [('Method 1', npv_1), ('Method 2', npv_2), ('Method 3', npv_3), ('Method 4', npv_4)]
for name, value in methods:
    diff = abs(value - excel_npv_opex)
    if diff < 0.01:
        print(f"\n✓ {name} MATCHES Excel! (diff = {diff:.4f})")

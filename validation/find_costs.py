import openpyxl

wb = openpyxl.load_workbook('oa_hybrid.xlsx', data_only=True)
ws = wb['LCOE']

print("Scanning for cost parameters...")
for i in range(1, 50):
    a_val = ws.cell(i, 1).value
    b_val = ws.cell(i, 2).value   
    c_val = ws.cell(i, 3).value
    d_val = ws.cell(i, 4).value
    
    if a_val:
        a_str = str(a_val).lower()
        if any(word in a_str for word in ['capex', 'opex', 'cost', 'tax', 'discount', 'equity', 'loan', 'depr']):
            print(f'Row {i:2d}: A={str(a_val):40s} B={b_val if b_val else "":<15} C={c_val if c_val else "":<15} D={d_val if d_val else ""}')

import openpyxl
import numpy as np

wb_formulas = openpyxl.load_workbook('oa_hybrid.xlsx')
wb_values = openpyxl.load_workbook('oa_hybrid.xlsx', data_only=True)
ws_f = wb_formulas['LCOE']
ws_v = wb_values['LCOE']

print("="*80)
print("EXCEL LCOE FORMULA BREAKDOWN")
print("="*80)

# Formula: =(E4-(E7+E17)*D9+E18+E5*(1-D9)-E6)/E12
print("\nLCOE Formula in B20: =(E4-(E7+E17)*D9+E18+E5*(1-D9)-E6)/E12")
print("\nComponent Values:")

e4 = ws_v['E4'].value  # PCI (Equity Capital Investment)
e5 = ws_v['E5'].value  # NPV of Annual Costs (OPEX)
e6 = ws_v['E6'].value  # NPV of Residual Value  
e7 = ws_v['E7'].value  # NPV of Depreciation
e12 = ws_v['E12'].value  # NPV of Annual Yield (Energy)
e17 = ws_v['E17'].value  # NPV of Interest
e18 = ws_v['E18'].value  # NPV of Loan Payments
d9 = ws_v['D9'].value  # Tax Rate

print(f"E4  (PCI - Equity):        ₹{e4:.2f} Crore")
print(f"E5  (NPV of OPEX):         ₹{e5:.2f} Crore")
print(f"E6  (NPV of Residual):     ₹{e6:.2f} Crore")
print(f"E7  (NPV of Depreciation): ₹{e7:.2f} Crore")
print(f"E12 (NPV of Energy):       {e12:.2f} (units?)")
print(f"E17 (NPV of Interest):     ₹{e17:.2f} Crore")
print(f"E18 (NPV of Loan Payment): ₹{e18:.2f} Crore")
print(f"D9  (Tax Rate):            {d9:.4f}")

print("\nCalculating LCOE:")
print(f"  Numerator components:")
print(f"    E4 (Equity):                   {e4:.2f}")
print(f"    -(E7+E17)*D9 (Tax Shield):    -{(e7+e17)*d9:.2f}")
print(f"    E18 (Loan Payments):           {e18:.2f}")
print(f"    E5*(1-D9) (OPEX after tax):    {e5*(1-d9):.2f}")
print(f"    -E6 (Residual):               -{e6:.2f}")

numerator = e4 - (e7+e17)*d9 + e18 + e5*(1-d9) - e6
denominator = e12

print(f"\n  Total Numerator:  {numerator:.2f}")
print(f"  Total Denominator: {denominator:.2f}")
print(f"  LCOE = {numerator:.2f} / {denominator:.2f} = {numerator/denominator:.10f}")

# Check what units are needed
lcoe_calculated = numerator / denominator
lcoe_actual = ws_v['B20'].value

print(f"\n  Calculated LCOE:  {lcoe_calculated:.10f}")
print(f"  Excel LCOE:       {lcoe_actual:.10f}")
print(f"  Match: {'YES' if abs(lcoe_calculated - lcoe_actual) < 0.0001 else 'NO'}")

# Understand the energy units
print("\n" + "="*80)
print("ENERGY / YIELD DATA")
print("="*80)

print(f"\nE12 (NPV of Energy): {e12:.2f}")
print(f"This should be in Crore kWh or similar units")

# Extract year 1 yield
h12 = ws_v['H12'].value
print(f"\nYear 1 Yield (H12): {h12:.2f}")

# Check row 12 across all years
print("\nRow 12 (Yield) across years:")
yields = []
for col in range(8, 33):  # H through AF (25 years)
    val = ws_v.cell(12, col).value
    if val:
        yields.append(val)
        if col < 11:  # Show first 3 years
            print(f"  Year {col-7}: {val:.2f}")

if len(yields) > 0:
    print(f"\n  Total years with data: {len(yields)}")
    print(f"  Year 1-3 sum: {sum(yields[:3]):.2f}")
    
    # Calculate NPV manually
    discount_rate = ws_v['D11'].value
    npv_energy_manual = sum(y / (1 + discount_rate)**n for n, y in enumerate(yields, 1))
    print(f"\n  Manual NPV calculation: {npv_energy_manual:.2f}")
    print(f"  Excel NPV (E12):        {e12:.2f}")
    print(f"  Match: {'YES' if abs(npv_energy_manual - e12) < 0.01 else 'NO'}")

# Check if yield is in Crore kWh
print("\n" + "="*80)
print("ENERGY UNIT CHECK")
print("="*80)

# We know from simulation: 1,915,710.47 MWh delivered in year 1
simulated_year1_mwh = 1915710.47
simulated_year1_kwh = simulated_year1_mwh * 1000
simulated_year1_crore_kwh = simulated_year1_kwh / 1e7

print(f"From Python simulation:")
print(f"  Year 1 delivered: {simulated_year1_mwh:,.2f} MWh")
print(f"  Year 1 delivered: {simulated_year1_kwh:,.2f} kWh")  
print(f"  Year 1 delivered: {simulated_year1_crore_kwh:.2f} Crore kWh")

print(f"\nFrom Excel (H12): {h12:.2f}")
print(f"Unit appears to be: {'Crore kWh' if abs(h12 - simulated_year1_crore_kwh) < 1 else 'Unknown'}")

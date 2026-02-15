import openpyxl

wb = openpyxl.load_workbook('oa_hybrid.xlsx', data_only=True)
ws = wb['Summary']
ws_hd = wb[' Hourly Data']
ws_lcoe = wb['LCOE']

print("="*70)
print("SYSTEM CAPACITIES")
print("="*70)
print(f"Solar AC:     {ws['B14'].value} MW")
print(f"Solar DC:     {ws['B15'].value} MW")
print(f"DC/AC Ratio:  {ws['B15'].value / ws['B14'].value if ws['B14'].value else 0:.2f}")
print(f"Wind WTGs:    {ws['B21'].value if ws['B21'].value else 0} (0 for this case)")
print(f"BESS Power:   {ws_hd['B5'].value} MW")
print(f"BESS Energy:  {ws_hd['B8'].value} MWh")

print("\n" + "="*70)
print("COSTS FROM EXCEL LCOE SHEET")
print("="*70)
print(f"Project Cost: {ws_lcoe['D3'].value:.2f} Crore")
print(f"Annual OPEX:  {ws_lcoe['D5'].value:.2f} Crore")
print(f"Tax Rate:     {ws_lcoe['D9'].value}")
print(f"Discount Rate:{ws_lcoe['D11'].value}")
print(f"Equity %:     {ws_lcoe['C4'].value * 100:.1f}%")
print(f"Loan %:       {ws_lcoe['B15'].value * 100:.1f}%")
print(f"Loan Term:    {ws_lcoe['B16'].value} years")

# Try to find cost breakdown
print("\n" + "="*70)
print("Looking for cost breakdown...")
print("="*70)
for i in range(20, 35):
    a = ws_lcoe.cell(i, 1).value
    b = ws_lcoe.cell(i, 2).value
    if a and b:
        print(f"Row {i}: {a:40s} = {b}")
